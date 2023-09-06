#!/usr/bin/env python
# -*- coding: UTF-8 -*-
"""
@Project ：adi-datatool 
@File    ：dbutils.py
@Author  ：LvYong
@Date    ：2022/3/3 13:33 
"""
import pymysql
import cv2
import numpy as np
import pandas as pd
from typing import List, Union, Dict, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Font
from openpyxl.styles.colors import BLUE
from tqdm import tqdm
from copy import deepcopy
from PIL import Image as PImage
from collections import defaultdict
from .comm import *
from .feature import *


__all__ = ['connect_database', 'query_production', 'query_image', 'save_json']


def connect_database(host='192.168.2.30', user="root", passwd="123456",
                     database='backplanemysql') -> pymysql.Connection:
    """
        连接数据库

    :param host: 数据库服务器地址
    :param user: 数据库服务器用户
    :param passwd: 数据库服务器登录密码
    :param database: 数据库名称
    :return: 数据库连接对象
    """
    try:
        return pymysql.connect(host=host, user=user, passwd=passwd, database=database)
    except Exception as ex:
        raise OSError(f'连接失败！{ex}')


def query_production(connection: pymysql.Connection, product_id, min_num=10) -> Dict:
    """
        查询指定产品生产信息

    :param connection: 数据库链接
    :param product_id: 产品ID, 即产品类型
    :param min_num: 每批次至少生产的产品数目
    :return: 产品批次ID, 产品数目, 生产时间
    """

    assert isinstance(connection, pymysql.Connection)
    cursor = connection.cursor()

    sql = f"select * from produce where product_id={product_id} and total >= {min_num};"
    cursor.execute(sql)
    product_infos = cursor.fetchall()
    cursor.close()
    batch_infos = {
        info[0]: {
            'batch_id': info[0],
            'product_num': info[2],
            'time': f'{info[6]}~{info[7]}'
        } for info in product_infos
    }

    return batch_infos


def query_image(connection: pymysql.Connection, batch_ids: Union[int, List[int]], save_type='json', save_path=None,
                save_image=False, save_thumb=True, save_ok_image=False, image_folders: Optional[List[str]] = None,
                QC_data: Optional[str] = None, retrain_filtered=True, split_thresh=3000):

    """
        查询产品图像信息

    :param connection: 数据库链接
    :param batch_ids: 查询产品的批次信息
    :param save_type: 信息保存类型，['json', 'excel', 'all']
    :param save_path: 信息保存路径，默认为系统桌面
    :param save_image: 是否获取图像并保存，默认False
    :param save_thumb: 是否保存缩略图，默认False
    :param save_ok_image: 是否保存OK原始图，默认False
    :param image_folders: 原始图像访问地址
    :param QC_data: 缺陷复判数据，json文件
    :param retrain_filtered: 查询结果是否包含已过滤的缺陷
    :param split_thresh: 查询图像数目大于阈值时，按照类别单独保存数据；反之，查询结果保存到一起. 默认3000
    :return: None
    """

    assert isinstance(connection, pymysql.Connection)
    assert save_type in ('json', 'excel', 'all')
    if isinstance(batch_ids, int):
        batch_sql = f'produce.produce_id={batch_ids}'
    elif isinstance(batch_ids, list) and len(batch_ids):
        if len(batch_ids) > 1:
            batch_sql = f'produce.produce_id in {tuple(batch_ids)}'
        else:
            batch_sql = f'produce.produce_id={batch_ids[0]}'
    else:
        raise ValueError(f"输入查询批次参数[batch_ids]格式错误:{batch_ids}")
    if save_image and not image_folders:
        raise ValueError("请配置图像原始存储地址")
    if QC_data and not QC_data.endswith('.json'):
        raise ValueError("缺陷复判数据为json格式文件的路径")
    if save_path is None:  # 捕捉当前目录的家目录，并在桌面上新建保存目录
        save_path = create_dir(os.path.join(os.path.expanduser('~'), 'Desktop', f'{get_cur_time("%m%d%H")}'))

    # 查询数据
    print(f'批次信息-{batch_ids},数据查询中...')
    cursor = connection.cursor()
    # soft_judge枚举值 0：阈值过滤, 1：roi过滤, 2：NG, 4：疑似
    sql = f"""Select image.image_path, image.work_no, image.camera_no, image.image_no, image.create_time, 
        dft.type, dft.score, dft.area, dft.model_d, dft.circularity, dft.gray, 
        dft.f_mark_c, dft.f_mark_r, dft.soft_judge, dft.mask_point_c, dft.mask_point_r, dft.width, dft.length,dft.filter
        from image, image_defect as dft 
        where image.id = dft.image_id 
        
        and dft.soft_judge in {'(0, 1, 2, 4)' if retrain_filtered else '(0, 2, 4)'}
        and image.produce_detail_id in (select produce.id from produce_detail as produce where {batch_sql})
        """
    cursor.execute(sql)
    defect_infos = cursor.fetchall()
    cursor.close()
    print(f'查询完成-{len(defect_infos)}条缺陷,数据解析中...')

    cursor = connection.cursor()
    sql = f"""Select image.image_path, image.work_no, image.camera_no, image.image_no, image.create_time
                from image
                where image.defect_nums  = 0
                and image.produce_detail_id in (select produce.id from produce_detail as produce where {batch_sql})
                """
    cursor.execute(sql)
    ok_image_infos = cursor.fetchall()
    cursor.close()
    print(f'查询完成-{len(ok_image_infos)}条OK图,数据解析中...')

    # 解析数据方法
    def _parse_sql_info(sql_info):
        _img_path, _path_infos = f'{sql_info[0]}.bmp', sql_info[0].split('/')
        _batch_id, _product_No, _name = _path_infos[-5], _path_infos[-4], _path_infos[-1]
        # _data_key = f'{_batch_id}_{_product_No}-{_name}.bmp'
        _data_key = f'{_product_No}-{_name}.jpg'
        _img_path = _img_path.replace('NG', "ORIG").replace('OK', "ORIG").replace('Suspect', "ORIG")
        _img_info = {
            'filename': _data_key,
            'sample': {
                'btNo': int(_batch_id),
                'pdtNo': int(_product_No),
                'workNo': sql_info[1],
                'camNo': sql_info[2],
                'imgNo': sql_info[3],
                'time': sql_info[4]
            },
            'regions': []
        }
        return _img_path, _batch_id, _product_No, _data_key, _img_info

    defect_data, batch_infos, image_infos, product_infos = {}, [], {}, []
    # 解析NG数据
    for defect_info in defect_infos:
        img_path, batch_id, product_No, data_key, img_info = _parse_sql_info(defect_info)
        batch_infos.append(batch_id)
        product_infos.append(product_No)
        if data_key not in defect_data:
            image_infos[data_key] = img_path
            img_info['regions'].append(gen_region(defect_info, offset=5))
            defect_data[data_key] = img_info
        else:
            defect_data[data_key]['regions'].append(gen_region(defect_info, offset=5,
                                                               region_idx=len(defect_data[data_key]['regions'])))
    # 解析OK数据
    for ok_image_info in ok_image_infos:
        img_path, batch_id, product_No, data_key, img_info = _parse_sql_info(ok_image_info)
        batch_infos.append(batch_id)
        product_infos.append(product_No)
        if data_key not in defect_data:
            defect_data[data_key] = img_info
            if save_ok_image:
                image_infos[data_key] = img_path

    print(f'解析完成,共{len(set(product_infos))}个产品,{len(defect_data)}张图像!')

    # 复判数据
    qc_check(defect_data, QC_data, inplace=True)

    # 保存图像
    thumbs = {}
    if save_image:
        thumbs = get_image(image_infos, defect_data, image_folders, save_path,
                           save_thumb=save_thumb, thumbnail=save_type in ['excel', 'all'])

    # 特征计算
    feature = get_feature()
    feature_map = list(get_feature_map(feature).values())
    with tqdm(defect_data.items(), desc='计算特征', ncols=80) as feature_pro:
        for k, dft_info in feature_pro:
            img_file = os.path.join(save_path, 'image', k.split('_')[0], k)
            src_img = cv2.imdecode(np.fromfile(img_file, dtype=np.uint8), 0) if os.path.exists(img_file) else None
            feature_data = feature.get_image_feature(src_img, dft_info['regions'])
            for d_idx, f_data in enumerate(feature_data):
                for f_idx, d in enumerate(f_data):
                    dft_info['regions'][d_idx]['region_attributes'][feature_map[f_idx]] = round(d, 4)

    # 保存数据
    if len(defect_data) <= split_thresh:
        # 单文件保存数据
        save_data(defect_data, thumbs, save_type, save_path, f'batch_{"_".join(set(batch_infos))}')
    else:
        # 按类别分别进行保存数据
        dft_class_data = defaultdict(dict)
        for k, v in defect_data.items():
            cls_regions = defaultdict(list)
            for region in v["regions"]:
                lb = region['region_attributes']['preLb']
                cls_regions[lb].append(deepcopy(region))
            for cls, regions in cls_regions.items():
                dft_data = deepcopy(v)
                dft_data['regions'] = regions
                dft_class_data[cls][k] = dft_data

        for k, v in dft_class_data.items():
            save_name = f'batch_{"_".join(set(batch_infos))}_{k}'
            save_data(v, thumbs, save_type, save_path, save_name)


def gen_region(data, offset=5, region_idx=0):
    """
        构建region字段

    :param data: 数据
    :param offset: 索引偏移量
    :param region_idx: 序号
    :return: dict
    """

    pt_xs = [int(p) for p in str(data[9 + offset]).split(',')] if data[9 + offset] else []
    pt_ys = [int(p) for p in str(data[10 + offset]).split(',')] if data[10 + offset] else []
    region = {
        'region_attributes': {
            # 推理标签
            'regions': int(data[0 + offset]),
            'preLb': int(data[0 + offset]),
            # QC_judge枚举值
            # -1:未知(QC没给出复判信息), 0:over, 1: NG
            'qcJdg': 0,  # 未复判前，所有检出都视为过检
            # 推理得分
            'score': round(data[1 + offset], 4),
            # 推理面积
            'preArea': round(data[2 + offset], 4),
            # 缺陷直径
            'dia': round(data[3 + offset], 4),
            # 缺陷圆度circularity
            'cir': round(data[4 + offset], 4),
            # 缺陷灰度差
            'gryDif': round(data[5 + offset], 2),
            # 缺陷灰度均值
            'gryMean': round(data[11 + offset], 2),
            # 缺陷标准差
            'gryStd': round(data[12 + offset], 2),
            # 缺陷中心点
            'ctrX': int(data[6 + offset]),
            'ctrY': int(data[7 + offset]),
            # soft_judge枚举值
            # 0：阈值过滤, 1：roi过滤, 2：NG, 4：疑似
            'softJdg': int(data[8 + offset]),
            # 当前缺陷在所有缺陷中的索引序号，便于按照类别划分缺陷后数据的查询
            'rIdx': region_idx,
            # 人工标签
            'mrkLb': -1,
            # 检测情况标签
            # 1：屏蔽区过滤 2：与1同情况 3：分类模型结果
            # 4：判定参数过滤 5：算法灰尘过滤（二次擦拭起作用） 6：规则（NGrule.ini） 7：脏污模型结果
            'filter-soft': round(data[13 + offset], 2)
        },
        'filter_attributes': {
            'roi': 1 if int(data[8 + offset]) == 1 else 0,
            'filter': '',
        },
        'shape_attributes': {
            'all_points_x': pt_xs,
            'all_points_y': pt_ys,
            'box': [min(pt_xs), min(pt_ys), max(pt_xs), max(pt_ys)] if len(pt_xs) else []
        }
    }
    return region


def qc_check(defect_data, qc_data, inplace=False):
    """
        根据QC复判标准，更新缺陷的qcJdg属性

    :param defect_data: 缺陷数据
    :param qc_data: QC复判数据
    :param inplace: 如果False,返回defect_data的深度复制数据. 否则, 直接修改defect_data属性
    :return: None
    """

    if not inplace:
        defect_data = deepcopy(defect_data)

    if qc_data and os.path.exists(qc_data):
        check_data = json.load(open(qc_data))
        for k, v in defect_data.items():
            if k in check_data:
                for region in v['regions']:
                    judge, x, y = 0, region['region_attributes']['ctrX'], region['region_attributes']['ctrY']
                    for box in check_data[k]['regions']:
                        judge = box_contains_point(box['shape_attributes']['all_points_x'],
                                                   box['shape_attributes']['all_points_y'],
                                                   (x, y))
                        if judge == 1:
                            break
                    region['region_attributes']['qcJdg'] = judge
    elif qc_data:
        print('缺陷复判数据不存在！')

    return defect_data


def save_data(defect_data: dict, thumb_data: dict, save_type, save_path, save_name):
    """
        保存缺陷信息

    :param defect_data: 缺陷数据
    :param thumb_data: 缩略图数据
    :param save_type: 信息保存类型，['json', 'excel', 'all']
    :param save_path: 信息保存路径，默认为系统桌面
    :param save_name: 保存名称
    :return: None
    """

    if save_type in ['json', 'all']:
        save_json(defect_data, save_path, save_name)
    if save_type in ['excel', 'all']:
        dft_data, data = list(defect_data.values()), []
        with tqdm(range(len(dft_data) + 1), desc=f'保存Excel[{save_name}]数据', ncols=80) as dft_pro:
            for idx in dft_pro:
                if idx != len(dft_data):
                    if len(dft_data[idx]['regions']):
                        data.append(pd.json_normalize(dft_data[idx],
                                                      record_path='regions',
                                                      meta=['filename',
                                                            ['sample', 'workNo'],
                                                            ['sample', 'camNo'],
                                                            ['sample', 'imgNo']],
                                                      errors='ignore'
                                                      ).drop(['shape_attributes.all_points_x',
                                                              'shape_attributes.all_points_y',
                                                              'shape_attributes.box',
                                                              'region_attributes.regions'],
                                                             axis=1))
                else:
                    data = pd.concat(data, ignore_index=True)
                    save_excel(data, save_path, save_name, insert_thumbs=thumb_data, freeze_panes='F')
                    dft_pro.refresh()


def get_image(image_infos, defect_infos, image_folders, save_path, save_thumb=False, thumbnail=False) -> dict:
    """
        从数据库获取图像并保存

    :param image_infos: 图像信息
    :param defect_infos: 缺陷信息
    :param image_folders: 原始图像文件夹
    :param save_path: 保存路径
    :param save_thumb: 是否保存缩略图
    :param thumbnail: 是否输出缩略图
    :return: 绘制小图
    """

    dft_thumbs = {}
    with tqdm(image_infos.items(), desc='保存图像', ncols=80, delay=0.05) as img_pro:
        for image_name, image_path in img_pro:
            img_save_path = create_dir(os.path.join(save_path, 'image', image_name.split('_')[0]))
            dst_file = os.path.join(img_save_path, image_name)
            if not os.path.exists(dst_file):
                for image_folder in image_folders:
                    image_file = os.path.join(image_folder, image_path)
                    if os.path.exists(image_file):
                        shutil.copy(image_file, dst_file)
                        break
            if save_thumb:
                dft_thumbs[image_name] = draw_defect(dst_file, defect_infos[image_name], img_save_path,
                                                     compared=False, save_format='.bmp', thumbnail=thumbnail)
            draw_defect(dst_file, defect_infos[image_name], img_save_path,
                        compared=False, save_format='.jpg', draw_single=False, thickness=10)
    return dft_thumbs


def draw_defect(image_file: str, dft_data: dict, save_folder, compared=True, draw_label=False, draw_single=True,
                color=(255, 0, 0), thickness=1, save_format='.jpg', thumbnail=False):
    """
        可视化缺陷，将标注json信息绘制到图像上。

    :param image_file: 图像文件
    :param dft_data: 缺陷信息
    :param save_folder: 保存路径
    :param compared: 是否同时输出对照图（未绘制任何标注信息的图像）
    :param draw_label: 是否同时绘制缺陷标签信息
    :param draw_single: 是否将每个缺陷作为独立个体绘制在图像上
    :param color: 缺陷绘制颜色
    :param thickness: 绘制缺陷边界线条宽度，≥1
    :param save_format: 保存格式，查看局部图时，建议为.bmp格式；查看整图时，建议.jpg格式
    :param thumbnail: 是否输出缩略图
    :return: 绘制结果保存路径
    """

    thumbs = []
    if not os.path.exists(image_file) or not dft_data:
        return thumbs

    full_name = dft_data['filename']
    regions = dft_data['regions']
    src_img = cv2.imdecode(np.fromfile(image_file, dtype=np.uint8), cv2.IMREAD_COLOR)   # -BGR
    comparison = src_img.copy() if compared else None
    img_size = src_img.shape[:2]

    for idx, region in enumerate(regions):
        try:
            region_label = int(region['region_attributes']['regions'])
        except KeyError:
            region_label = 1

        if draw_single:
            save_temp = create_dir(os.path.join(save_folder, 'thumb', str(region_label)))
            save_file = os.path.join(save_temp, f'{os.path.splitext(full_name)[0]}_{idx}{save_format}')
            if os.path.exists(save_file) and thumbnail:
                thumbs.append(Image(save_file))
                continue

            drawn_img = draw_via_region(src_img, region, color, thickness, draw_label, inplace=False)
            # 每个缺陷作为独立结果输出
            box, anchor = get_via_region_boundary(region, 15, img_size)
            df_cmp_img = src_img[box[1]:box[3], box[0]:box[2]]
            df_mrk_img = drawn_img[box[1]:box[3], box[0]:box[2]]

            if compared:
                border = 2
                bg_img = PImage.new('RGB',
                                    ((box[2] - box[0]) * 2 + border * 3, (box[3] - box[1]) + border * 2),
                                    "white")
                bg_img.paste(PImage.fromarray(df_cmp_img), (border, border))
                bg_img.paste(PImage.fromarray(df_mrk_img), (box[2] - box[0] + border * 2, border))
                bg_img = PImage.fromarray(cv2.cvtColor(np.asarray(bg_img), cv2.COLOR_RGB2BGR))
                bg_img.save(save_file)
                if thumbnail:
                    thumbs.append(Image(bg_img))
            else:
                cv2.imencode(save_format, df_mrk_img)[1].tofile(save_file)
                if thumbnail:
                    df_mrk_img = PImage.fromarray(cv2.cvtColor(df_mrk_img, cv2.COLOR_BGR2RGB))
                    thumbs.append(Image(df_mrk_img))
        else:
            save_temp = create_dir(os.path.join(save_folder, 'large'))
            save_file = os.path.join(save_temp, os.path.splitext(full_name)[0] + save_format)
            if os.path.exists(save_file):
                break

            draw_via_region(src_img, region, color, thickness, draw_label)

            # 绘制所有缺陷作为输出
            if idx + 1 == len(regions):
                if compared:
                    border = 2
                    bg_img = PImage.new('RGB', (img_size[1] * 2 + border * 3, img_size[0] + border * 2), "#FFFFFF")
                    bg_img.paste(PImage.fromarray(comparison), (border, border))
                    bg_img.paste(PImage.fromarray(src_img), (img_size[1] + border * 2, border))
                    bg_img = PImage.fromarray(cv2.cvtColor(np.asarray(bg_img), cv2.COLOR_RGB2BGR))
                    bg_img.save(save_file)
                else:
                    cv2.imencode(save_format, src_img)[1].tofile(save_file)

    del src_img, comparison
    return thumbs


def draw_via_region(image, region, color=(255, 0, 0), thickness=1, draw_label=False, inplace=True):
    """
        绘制via区域信息，直接更新输入图像

    :param image: 图像，进过绘制或发生变化
    :param region: via的区域字段信息
    :param color: 区域颜色
    :param thickness: 区域边界宽度
    :param draw_label: 是否同时绘制缺陷标签信息
    :param inplace: 是否执行原地操作
    :return: 绘制后的图像
    """

    region_shape = region['shape_attributes']['name'] if 'name' in region['shape_attributes'] else 'polygon'
    try:
        region_label = int(region['region_attributes']['regions'])
    except KeyError:
        region_label = 1

    drawn_image = image if inplace else image.copy()

    # 绘制标签
    if region_label and draw_label:
        font_face, font_scale = cv2.FONT_HERSHEY_SIMPLEX, 0.5
        # 计算文本的宽高，baseLine
        text_size, base_line = cv2.getTextSize(str(region_label), font_face, font_scale, thickness)
        box, anchor = get_via_region_boundary(region, 5, drawn_image.shape[:2])
        if anchor[1] < text_size[1] + base_line:
            anchor[1] = text_size[1] + base_line
            anchor[0] += base_line
        if anchor[0] < text_size[0] + base_line:
            anchor[0] = text_size[0] + base_line
            anchor[1] += base_line
        cv2.putText(drawn_image, str(region_label), (anchor[0], anchor[1]),
                    font_face, font_scale, color, thickness)

    # 绘制区域
    if region_shape == "circle":
        cx = region['shape_attributes']['cx']
        cy = region['shape_attributes']['cy']
        r = region['shape_attributes']['r']
        cv2.circle(drawn_image, (cx, cy), int(r), color=color, thickness=thickness)
    elif region_shape == "rect":
        left = region['shape_attributes']['x']
        top = region['shape_attributes']['y']
        right = region['shape_attributes']['x'] + region['shape_attributes']['width']
        bottom = region['shape_attributes']['y'] + region['shape_attributes']['height']
        cv2.rectangle(drawn_image, (left, top), (right, bottom), color=color, thickness=thickness)
    elif region_shape == "polygon":
        ptx = region['shape_attributes']['all_points_x']
        pty = region['shape_attributes']['all_points_y']
        ptx0, pty0 = ptx[0], pty[0]
        _ptx, _pty = [ptx0], [pty0]
        for i in range(1, len(ptx)):
            if ptx[i] - ptx0 < 2 and pty[i] - pty0 < 2:
                _ptx.append(ptx[i])
                _pty.append(pty[i])
            else:
                points = np.dstack((_ptx, _pty))
                cv2.polylines(drawn_image, pts=[points], isClosed=True, color=color, thickness=thickness)
                _ptx, _pty = [ptx[i]], [pty[i]]
            ptx0, pty0 = ptx[i], pty[i]
        if len(_ptx) > 3:
            points = np.dstack((_ptx, _pty))
            cv2.polylines(drawn_image, pts=[points], isClosed=True, color=color, thickness=thickness)
    elif region_shape == "ellipse":
        cx = int(region['shape_attributes']['cx'])
        cy = int(region['shape_attributes']['cy'])
        l_x = int(region['shape_attributes']['rx'])
        s_y = int(region['shape_attributes']['ry'])
        theta = int(region['shape_attributes']['theta'])
        cv2.ellipse(drawn_image, (cx, cy), (l_x, s_y), theta, theta, 360, color=color, thickness=thickness)
    else:
        raise ValueError(f'不支持解析[{region_shape}]区域!!!')

    return drawn_image


def get_via_region_boundary(region, extension=0, limit_size=None):
    """
        获取区域的外接矩形

    :param region: 区域信息
    :param extension: 外扩像素数目
    :param limit_size: 限制区域，一般为图像大小(rows, cols)
    :return: 区域外接矩形信息 (left, top, right, bottom)，区域边界离外接矩形最近的点 anchor
    """

    region_shape = region['shape_attributes']['name'] if 'name' in region['shape_attributes'] else 'polygon'
    anchor = [0, 0]

    if region_shape == "circle":
        cx = region['shape_attributes']['cx']
        cy = region['shape_attributes']['cy']
        r = region['shape_attributes']['r']
        box = (cx - r, cy - r, cx + r, cy + r)
    elif region_shape == "rect":
        left = region['shape_attributes']['x']
        top = region['shape_attributes']['y']
        right = region['shape_attributes']['x'] + region['shape_attributes']['width']
        bottom = region['shape_attributes']['y'] + region['shape_attributes']['height']
        box = (left, top, right, bottom)
    elif region_shape == "polygon":
        ptx = region['shape_attributes']['all_points_x']
        pty = region['shape_attributes']['all_points_y']
        min_y = min(pty)
        box = (min(ptx), min_y, max(ptx), max(pty))
        anchor = [ptx[pty.index(min_y)], min_y]
        anchor[0] = 0 if anchor[0] - extension < 0 else anchor[0] - extension
        anchor[1] = 0 if anchor[1] - extension < 0 else anchor[1] - extension
    else:
        raise ValueError(f'不支持解析[{region_shape}]区域!!!')

    box = [box[0] - extension, box[1] - extension, box[2] + extension, box[3] + extension]
    box[0] = 0 if box[0] < 0 else box[0]
    box[1] = 0 if box[1] < 0 else box[1]

    if limit_size:
        box[2] = limit_size[1] if box[2] > limit_size[1] else box[2]
        box[3] = limit_size[0] if box[3] > limit_size[0] else box[3]

    if region_shape != "polygon":
        anchor = [box[0], box[1]]

    return tuple(box), anchor


def save_excel(data, path, excel_name: str = None, sheet_name: str = None,
               freeze_panes: Union[int, str] = None, insert_thumbs: Optional[dict] = None) -> None:
    """
        保存模型性能指标到excel中

    :param data: 需保存的数据
    :param path: 保存路径
    :param excel_name: excel文件名称，默认为 ‘新建 Microsoft Excel 工作表’
    :param sheet_name: excel表名称
    :param freeze_panes: 冻结第二行第几列窗口（默认不冻结窗口）, (1 -> 'A')
    :param insert_thumbs: 插入的缩略图数据
    :return:
    """

    assert isinstance(data, pd.DataFrame)
    if not len(data):
        return

    excel_name = excel_name if excel_name else '新建 Microsoft Excel 工作表'
    file_name = excel_name if excel_name.endswith('.xlsx') else f'{excel_name}.xlsx'
    file = os.path.join(path, file_name)

    def _save(f):
        work_book = None
        try:
            if os.path.exists(f):
                work_book = load_workbook(f)
            with pd.ExcelWriter(f, engine='openpyxl') as writer:
                if work_book:
                    writer.book = work_book
                sh_name = f'Sheet_{get_cur_time()}' if sheet_name is None else sheet_name
                data.to_excel(excel_writer=writer, sheet_name=sh_name, index=True)
                data_rows, data_cols = data.shape

                # 冻结窗口
                worksheet = writer.sheets[sh_name]
                worksheet.cell(1, 1, 'idx')
                if freeze_panes:
                    col_letter = get_column_letter(freeze_panes) if isinstance(freeze_panes, int) else freeze_panes
                    worksheet.freeze_panes = f'{col_letter}2'
                worksheet['D2'].alignment = Alignment(horizontal='center', vertical='center')

                # 插入缩略图
                if insert_thumbs:
                    # 获取图像最大的尺寸并限定图像尺寸
                    max_width, max_height = 10, 10
                    thresh_size = 100
                    thumb_rates = {}
                    for k, v in insert_thumbs.items():
                        thumb_rates[k] = []
                        for thumb in v:
                            if thumb.width > thresh_size or thumb.height > thresh_size:
                                thumb_rates[k].append(thresh_size / max(thumb.width, thumb.height))
                                if thumb.width > thresh_size:
                                    max_width = thresh_size
                                if thumb.height > thresh_size:
                                    max_height = thresh_size
                            else:
                                thumb_rates[k].append(1)
                                if thumb.width > max_width:
                                    max_width = thumb.width
                                if thumb.height > max_height:
                                    max_height = thumb.height

                    # 考虑增加Index列，所以需要偏移1列
                    img_letter = get_column_letter(data_cols + 2)
                    worksheet.column_dimensions[img_letter].width = max_width * 0.14
                    name_idx = data.columns.to_list().index('filename') + 1
                    region_idx = data.columns.to_list().index('region_attributes.rIdx') + 1
                    worksheet.cell(1, column_index_from_string(img_letter), 'image')
                    link_font = Font(name='宋体', underline='single', color=BLUE)
                    # 插入图像数据
                    for idx in range(data_rows):
                        row_idx = idx + 2
                        # 通过数字访问单元格时,行索引从1开始,列索引从0开始
                        filename = worksheet[row_idx][name_idx].value
                        # 原图名称单元格设定超链接
                        img_path = f'image\\{filename.split("_")[0]}\\large\\{filename}'
                        worksheet[row_idx][name_idx].hyperlink = img_path.replace('.bmp', '.jpg')
                        worksheet[row_idx][name_idx].font = link_font
                        # 插入图像
                        rIdx = worksheet[row_idx][region_idx].value
                        if filename in insert_thumbs and rIdx < len(insert_thumbs[filename]):
                            dft_thumb = insert_thumbs[filename][rIdx]
                            dft_rate = thumb_rates[filename][rIdx]
                            dft_thumb.width, dft_thumb.height = dft_thumb.width * dft_rate, dft_thumb.height * dft_rate
                            worksheet.row_dimensions[row_idx].height = max_height * 0.78
                            worksheet.add_image(dft_thumb, f'{img_letter}{row_idx}')
                        else:
                            print(rIdx)

                # 所有单元格居中
                name_idx = data.columns.to_list().index('filename') + 1
                name_letter = get_column_letter(name_idx + 1)  # 列的索引与字母转换之间差1
                worksheet.column_dimensions[name_letter].width = 20
                align = Alignment(horizontal='center', vertical='center')
                for c in range(1, data_cols + 3):
                    for cell in worksheet[get_column_letter(c)]:
                        cell.alignment = align

                # 更改列名称
                for c in range(0, data_cols + 2):
                    worksheet[1][c].value = str(worksheet[1][c].value).split('.')[-1]
        except Exception as e:
            raise e
        finally:
            if work_book:
                work_book.close()

    try:
        _save(file)
    except PermissionError:
        print(f'保存{file_name}异常！尝试重命名进行保存。')
        temp_file = file.replace('.xlsx', f'_{get_cur_time()}.xlsx')
        if os.path.exists(file):
            shutil.copy(file, temp_file)
        _save(temp_file)


# if __name__ == '__main__':

#     with connect_database(host='192.168.2.98') as db_connect:
#     # print(query_production(db_connect, 6))
#         query_image(db_connect,367, save_type='all', save_image=False, save_thumb=False, save_ok_image = False,
#                     retrain_filtered=True,
#                     split_thresh=10000000000000000,
#                     image_folders=[
#                         r"\\TEST03slave\MiddleFrame-A-OQC\01-project\x64\Release\Data\X12\Image"],
#                     QC_data="",
#                     save_path=r"D:\PAD\soft"
#                     )